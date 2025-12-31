from openpyxl import load_workbook

EXCEL_PATH = 'C://Users//HP//PycharmProjects//pythonProject//project2_orangehrm//test_data//test_data.xlsx'
SHEET_NAME = "testdata"

def read_data():
    wb = load_workbook(EXCEL_PATH)
    sheet = wb[SHEET_NAME]

    data = []
    for row in range(2, sheet.max_row + 1):
        username = sheet.cell(row=row, column=1).value
        password = sheet.cell(row=row, column=2).value
        data.append((username, password, row))

    return data


def write_result(row, result):
    wb = load_workbook(EXCEL_PATH)
    sheet = wb[SHEET_NAME]

    sheet.cell(row=row, column=3).value = result
    wb.save(EXCEL_PATH)




